from flask import Flask, render_template, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

def get_month_data(month_name):
    wb = load_workbook("milk_data.xlsx")
    sheet = wb.active
    data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        month, paid, days_in_month, days_absent, days_coming, amount = row[:6]
        if month and month.strip().upper() == month_name.upper():
            data = {
                "Month": month,
                "Paid": paid,
                "Days in Month": days_in_month,
                "Days Absent": days_absent,
                "Days Coming": days_coming,
                "Amount": amount
            }
            break
    wb.close()
    return data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/month/<month_name>')
def month_data(month_name):
    data = get_month_data(month_name)
    if data:
        return jsonify(data)
    else:
        return jsonify({"error": "No data found for this month"})

if __name__ == "__main__":
    app.run(debug=True)
